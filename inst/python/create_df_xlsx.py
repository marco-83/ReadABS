import xlrd
import itertools
from xlrd.sheet import ctype_text
import pandas as pd
import os
from ABStable_xlsx import *
from openpyxl import load_workbook

from textwrap import fill
import copy
import numpy as np
import operator
from itertools import permutations

# Tested spreadsheets:
# 63060do001_201805.xls
# table 90a key information, by states and territories, 2018 to 2019.xls
# 81670do001_201718.xls
# 34180do0001_201617_1.xls
# 41300_table7_2017-18.xls # TODO: need to make it not run second table, which are the standard errors
# 4329000006do003_2011.xls
#table 32a non-special schools by primary enrolment, 2004-2019.xls # TODO: need to make it not run 1st table, which is a pivot table
# TODO: also with table 32a need to make it include the extra metadata


def main_xlsx(excel_workbook, allowed_blank_rows=1, spreadsheet_type="other"):#, filter_tabs=True):
    xl_workbook = load_workbook(filename=excel_workbook)
    # if filter_tabs == "True":
    #     filter_tabs = True
    # elif filter_tabs == "False":
    #     filter_tabs = False

    xl_workbook, data_sheets = import_spreadsheet(excel_workbook=xl_workbook, filter_tabs=False)
    tables = define_table(xl_workbook, data_sheets, allowed_blank_rows, spreadsheet_type)

    for table in tables:
        # if table.sheet_name == "Table 4A.1":
        #     print(table)
        if table.row_descriptions == "Could not locate row headings":
            tables.remove(table)
            continue
        if not table.table_completed:
            tables.remove(table)
    if not tables:
        return pd.DataFrame(), pd.DataFrame()

    results, tables = create_dataframes(tables, xl_workbook, spreadsheet_type)
    result_info = table_info(tables)

    # print(results[1])
    # results[0].to_csv('dataframe0.csv')
    # results[1].to_csv('dataframe1.csv')
    #results[2].to_csv('dataframe2.csv')
    #results[3].to_csv('dataframe3.csv')
    #results[4].to_csv('dataframe4.csv')
    #results[5].to_csv('dataframe5.csv')
    return results, result_info


def table_info(tables):
    """ Save information about the extracted data from the spreadsheet """
    frames = []
    for i, table in enumerate(tables):
        #df = pd.DataFrame.from_records([s.to_dict() for s in table])
        #df = pd.DataFrame.from_dict(table.to_dict())
        #print(table.rows)
        df = pd.DataFrame(table.to_dict(), index=[0])
        df['Table'] = 'Table' + str(i+1)
        df = df[['Table'] + [col for col in df.columns if col != 'Table']]
        frames.append(df)

    result_info = pd.concat(frames).set_index('Table')

    return result_info


def create_dataframes(tables, xl_workbook, spreadsheet_type):
    # Initialise a list that will contain the datafames
    results = []
    t = 0
    tables_to_remove = []
    results_to_remove = []

    # for i, table in enumerate(tables):
    #     results.append(pd.DataFrame())
    #     results[t] = extract_data(table, xl_workbook, spreadsheet_type, df=results[t])
    #     column_headings, column_subheadings = extract_column_headings(xl_workbook, table, spreadsheet_type)
    #     results[t] = pivot_table(column_headings, column_subheadings, xl_workbook, table, spreadsheet_type,
    #                              df=results[t])
    #     t += 1


    for i, table in enumerate(tables):
        try:
            results.append(pd.DataFrame())
            results[t] = extract_data(table, xl_workbook, spreadsheet_type, df=results[t])
            column_headings, column_subheadings = extract_column_headings(xl_workbook, table, spreadsheet_type)
            results[t] = pivot_table(column_headings, column_subheadings, xl_workbook, table, spreadsheet_type,
                                     df=results[t])
        except (TypeError, KeyError, ValueError):
            tables_to_remove.append(table)
            results_to_remove.append(t)
            t += 1
        else:
            t += 1
    # Remove results and tables that have errors
    if tables_to_remove:
        for table in tables_to_remove:
            tables.remove(table)
        results = [i for j, i in enumerate(results) if j not in results_to_remove]

    return results, tables


def extract_data(table, xl_workbook, spreadsheet_type, df):
    sheet = xl_workbook.get_sheet_by_name(table.sheet_name)
    # print("number_format D7", xl_workbook.get_sheet_by_name('Employee jobs index')["D7"].number_format)
    # print("table.sheet_name", table.sheet_name)
    # print(xl_workbook.get_sheet_by_name('Employee jobs index').cell_value(rowx=6, colx=4))
    # Add data to the dataframe
    first_row = min(table.rows)
    first_col = min(table.cols)
    for r in table.rows:
        for c in table.cols:
            df.loc[r - first_row, c - first_col] = sheet.cell(row=r, column=c).value

    # Reset row indices so they go 0,1,2...nmain.py
    df = df.reset_index(drop=True)
    # Add row descriptions to the dataframe
    row_descriptions = table.row_descriptions
    indentation_levels = table.indentation_levels

    if table.table_type == "long format":
        # for i, col in enumerate(reversed(row_descriptions)):
        #     df.insert(loc=i, column='descriptor_col_' + str(i), value=['' for j in range(df.shape[0])])
        #     for r in table.rows:
        #         cell = sheet.cell(row=r, column=c).value
        #         if is_date_format(cell.number_format):
        #             df.loc[r - first_row, 'descriptor_col_' + str(i)] = pd.to_datetime(cell.value).strftime('%d/%m/%Y')
        #         else:
        #             df.loc[r - first_row, 'descriptor_col_' + str(i)] = cell.value
        row_headings = merged_data_row_headings_function(xl_workbook, sheet_name=table.sheet_name,
                                                         merged_data_rows=table.merged_meta_data_row_headings,
                                                         data_rows=table.rows,
                                                         row_descriptions=table.row_descriptions,
                                                         row_titles=table.row_titles,
                                                         top_header_row=table.top_header_row)
        df = df.join(row_headings)

    if table.table_type == "wide format":
        # Insert empty columns into dataframe, where the descriptions will go
        if table.columns_with_indentation:
            for c in table.columns_with_indentation:
                for idx, i in enumerate(indentation_levels[c]):
                    df.insert(loc=idx, column='descriptor_col_' + str(c) + str(i), value=['' for j in range(df.shape[0])])

            row = 0
            top_cell_row = min(table.rows)

            for r in table.rows:
                column = 0
                for c in table.columns_with_indentation:
                    for i in indentation_levels[c]:
                        cell_row = r
                        cell = sheet.cell(row=cell_row, column=c)
                        indentation_cell = int(sheet.cell(cell_row, c).alignment.indent)

                        while indentation_cell > i or cell.value is None:
                            cell_row -= 1
                            if cell_row < 1:
                                break
                            cell = sheet.cell(row=cell_row, column=c)
                            indentation_cell = int(sheet.cell(cell_row, c).alignment.indent)
                            if cell_row < top_cell_row:
                                top_cell_row = cell_row
                        if indentation_cell == i:
                            if is_date_format(cell.number_format):
                                df.loc[row, df.columns[column]] = pd.to_datetime(cell.value).strftime('%d/%m/%Y')
                            else:
                                df.loc[row, df.columns[column]] = cell.value
                        column += 1
                    row += 1

        other_columns = set(i for i in table.row_descriptions if i not in table.columns_with_indentation)
        if other_columns and spreadsheet_type != "Time series":
            row_headings = merged_data_row_headings_function(xl_workbook, sheet_name=table.sheet_name,
                                                             merged_data_rows=table.merged_meta_data_row_headings,
                                                             data_rows=table.rows,
                                                             row_descriptions=other_columns,
                                                             row_titles=table.row_titles,
                                                             top_header_row=table.top_header_row)
            df = df.join(row_headings)

    return df


def extract_column_headings(xl_workbook, table, spreadsheet_type):
    sheet = xl_workbook.get_sheet_by_name(table.sheet_name)

    if spreadsheet_type in ["Time series"]:
        column_headings = pd.DataFrame()
        for r, title in table.column_titles.items():
            for c in table.cols:
                cell = sheet.cell(row=r, column=c)
                row_position = c - min(table.cols)
                if is_date_format(cell.number_format):
                    column_headings.loc[row_position, title] = pd.to_datetime(cell.value).strftime('%d/%m/%Y')
                else:
                    column_headings.loc[row_position, title] = cell.value

        return column_headings, {}

    else:
        first_row = min(table.rows)
        if table.table_type == "long format":
            column_headings = []
            for c in list(table.cols) + table.row_descriptions:
                col_header = sheet.cell(row=first_row-1, column=c).value
                column_headings.append(col_header)
            return column_headings, {}
        # if it is wide format:
        column_headings = merged_data_function(xl_workbook, sheet_name=table.sheet_name,
                                               merged_data_cols=table.merged_meta_data, data_cols=table.cols,
                                               data_rows=table.rows, #rows=table.column_header_locations,
                                               extra_rows=table.extra_meta_data, spreadsheet_type=spreadsheet_type,
                                               column_header_locations=table.column_header_locations,
                                               last_row_in_sheet=table.last_row_in_sheet)

        column_subheadings = merged_data_subheadings_function(xl_workbook, sheet_name=table.sheet_name,
                                                              merged_data_cols=table.merged_meta_data,
                                                              data_cols=table.cols, data_rows=table.rows,
                                                              rows=table.column_header_locations,
                                                              extra_rows=table.extra_meta_data, top_row=table.top_row)

    return column_headings, column_subheadings


def pivot_table(column_headings, column_subheadings, xl_workbook, table, spreadsheet_type, df):
    """ Pivot the data """

    if table.table_type == "long format":
        df = df.rename(columns={0: "value"})
        # ensure value column is at the end
        df = df[[c for c in df if c not in ['value']] + ['value']]
        return df

    sheet = xl_workbook.get_sheet_by_name(table.sheet_name)
    data_cols = table.cols
    merged_meta_data = table.merged_meta_data

    # Add column subheadings
    df = df.join(column_subheadings)


    # id variables are the descriptor columns; value variables are the data columns
    id_variables = []
    value_variables = []
    column_number = 0
    for i in list(df.columns):
        if isinstance(i, int):
            df.rename(columns={i: column_number}, inplace=True)
            value_variables.append(column_number)
            column_number += 1
        else:
            id_variables.append(i)

    # Pivot longer
    df = df.melt(id_vars=id_variables, value_vars=value_variables, var_name="column_number")
    column_headings['column_number'] = range(0, len(column_headings))

    # Merge in column descriptions
    df = df.merge(column_headings, on='column_number', how='right')
    df = df.drop(columns=['column_number'])
    # if table.sheet_name == "Table_1":
    #     print("df", df)

    # Rename columns and re-order them
    if spreadsheet_type == 'Time series':
        df = df.rename(columns={"descriptor_col_0": "Date"})

    df.columns = df.columns.astype(str)

    cols = df.columns.tolist()

    cols1 = [i for i in cols if len(i) >= 14 if i[:14] == 'descriptor_col']
    cols2 = [i for i in cols if len(i) >= 8 if i[:8] == 'Col_desc']
    cols3 = [i for i in cols if i not in cols1 and i not in cols2 and i != 'value']

    cols_update = cols1 + cols2 + cols3
    cols_update.append('value')

    # Check if there are any more columns. There shouldn't be. But if there is, then add them
    cols_more = [i for i in cols_update if i not in cols]
    cols = cols_update + cols_more
    df = df[cols]

    return df

